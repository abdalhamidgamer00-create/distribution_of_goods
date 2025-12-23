"""Styling helpers for Excel formatting."""

from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd


def style_header_row(worksheet) -> None:
    """Style the header row with blue background and white text."""
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')


def adjust_column_widths(worksheet, dataframe: pd.DataFrame) -> None:
    """Adjust column widths based on content type."""
    width_map = {
        'code': 12, 'product_name': 40, 'quantity': 15, 
        'sender_balance': 15, 'receiver_balance': 15, 
        'target_branch': 15, 'transfer_type': 15
    }
    for column_index, column in enumerate(dataframe.columns, 1):
        column_letter = get_column_letter(column_index)
        worksheet.column_dimensions[column_letter].width = width_map.get(
            column, 12
        )


def add_borders(worksheet, last_row: int, num_columns: int) -> None:
    """Add thin borders to all cells."""
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    for row in worksheet.iter_rows(
        min_row=1, max_row=last_row, min_col=1, max_col=num_columns
    ):
        for cell in row:
            cell.border = thin_border
