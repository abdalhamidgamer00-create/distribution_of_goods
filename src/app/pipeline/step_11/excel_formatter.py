"""
Excel formatter for Step 11

Converts CSV files to Excel with conditional formatting for balance columns.
- Red shades: values close to 0 (low stock warning)
- Green shades: values far from 0 (healthy stock)
"""

import os
import pandas as pd
from typing import List, Dict
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from src.shared.utils.logging_utils import get_logger

logger = get_logger(__name__)


# =============================================================================
# PUBLIC API
# =============================================================================

def convert_to_excel_with_formatting(csv_files: List[Dict], excel_output_dir: str) -> List[str]:
    """Convert CSV files to Excel with conditional formatting."""
    return [path for path in (_process_single_csv(file_info, excel_output_dir) for file_info in csv_files) if path]


# =============================================================================
# CSV PROCESSING HELPERS
# =============================================================================

def _process_single_csv(file_info: dict, excel_output_dir: str) -> str:
    """Process a single CSV file and return Excel path or None."""
    csv_path = file_info.get('path')
    if not csv_path or not os.path.exists(csv_path):
        return None
    try:
        return _convert_single_file(csv_path, excel_output_dir)
    except Exception as error:
        logger.warning(f"Error converting {csv_path}: {error}")
        return None


def _convert_single_file(csv_path: str, excel_output_dir: str) -> str:
    """Convert a single CSV file to Excel with formatting."""
    dataframe = pd.read_csv(csv_path, encoding='utf-8-sig')
    if dataframe.empty:
        return None
    excel_path = _determine_output_path(csv_path, excel_output_dir)
    dataframe.to_excel(excel_path, index=False, engine='openpyxl')
    _apply_conditional_formatting(excel_path, dataframe)
    logger.debug(f"Generated Excel: {os.path.basename(excel_path)}")
    return excel_path


# =============================================================================
# OUTPUT PATH HELPERS
# =============================================================================

def _get_output_subdir(excel_output_dir: str, folder_name: str, grandparent_name: str) -> str:
    """Get output subdirectory based on folder structure."""
    if folder_name.startswith('to_'):
        return os.path.join(excel_output_dir, grandparent_name, folder_name)
    return os.path.join(excel_output_dir, folder_name)


def _determine_output_path(csv_path: str, excel_output_dir: str) -> str:
    """Determine the output path for Excel file based on CSV structure."""
    csv_directory = os.path.dirname(csv_path)
    parent_directory = os.path.dirname(csv_directory)
    
    output_subdir = _get_output_subdir(excel_output_dir, os.path.basename(csv_directory), os.path.basename(parent_directory))
    os.makedirs(output_subdir, exist_ok=True)
    
    return os.path.join(output_subdir, os.path.basename(csv_path).replace('.csv', '.xlsx'))


# =============================================================================
# HEADER STYLING HELPERS
# =============================================================================

def _style_header_row(worksheet) -> None:
    """Style the header row with blue background and white text."""
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')


# =============================================================================
# BALANCE COLUMN HELPERS
# =============================================================================

def _find_balance_columns(dataframe: pd.DataFrame) -> tuple:
    """Find indices of sender and receiver balance columns."""
    sender_column_index, receiver_column_index = None, None
    for column_index, column in enumerate(dataframe.columns, 1):
        if column == 'sender_balance':
            sender_column_index = column_index
        elif column == 'receiver_balance':
            receiver_column_index = column_index
    return sender_column_index, receiver_column_index


def _create_color_rule() -> ColorScaleRule:
    """Create color scale rule: Red (0) -> Yellow (15) -> Green (30+)."""
    return ColorScaleRule(
        start_type='num', start_value=0, start_color='FF0000',
        mid_type='num', mid_value=15, mid_color='FFFF00',
        end_type='num', end_value=30, end_color='00FF00'
    )


def _apply_color_scale_to_column(worksheet, column_index: int, last_row: int, color_rule) -> None:
    """Apply color scale formatting to a column."""
    column_letter = get_column_letter(column_index)
    worksheet.conditional_formatting.add(f"{column_letter}2:{column_letter}{last_row}", color_rule)


def _apply_balance_formatting(worksheet, dataframe: pd.DataFrame, last_row: int) -> None:
    """Apply color scale formatting to balance columns."""
    sender_column_index, receiver_column_index = _find_balance_columns(dataframe)
    color_rule = _create_color_rule()
    
    if sender_column_index:
        _apply_color_scale_to_column(worksheet, sender_column_index, last_row, color_rule)
    if receiver_column_index:
        _apply_color_scale_to_column(worksheet, receiver_column_index, last_row, color_rule)


# =============================================================================
# FORMATTING HELPERS
# =============================================================================

def _adjust_column_widths(worksheet, dataframe: pd.DataFrame) -> None:
    """Adjust column widths based on content type."""
    width_map = {
        'code': 12, 'product_name': 40, 'quantity': 15, 
        'sender_balance': 15, 'receiver_balance': 15, 
        'target_branch': 15, 'transfer_type': 15
    }
    for column_index, column in enumerate(dataframe.columns, 1):
        column_letter = get_column_letter(column_index)
        worksheet.column_dimensions[column_letter].width = width_map.get(column, 12)


def _add_borders(worksheet, last_row: int, num_columns: int) -> None:
    """Add thin borders to all cells."""
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    for row in worksheet.iter_rows(min_row=1, max_row=last_row, min_col=1, max_col=num_columns):
        for cell in row:
            cell.border = thin_border


def _apply_conditional_formatting(excel_path: str, dataframe: pd.DataFrame) -> None:
    """Apply conditional formatting to balance columns."""
    workbook = load_workbook(excel_path)
    worksheet = workbook.active
    last_row = len(dataframe) + 1
    _style_header_row(worksheet)
    _apply_balance_formatting(worksheet, dataframe, last_row)
    _adjust_column_widths(worksheet, dataframe)
    _add_borders(worksheet, last_row, len(dataframe.columns))
    worksheet.freeze_panes = 'A2'
    workbook.save(excel_path)
    workbook.close()
