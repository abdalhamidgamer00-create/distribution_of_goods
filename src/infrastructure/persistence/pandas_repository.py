"""Pandas implementation of the DataRepository interface."""

import os
import pandas as pd
from typing import List, Dict, Optional
from src.domain.models.entities import (
    Product, Branch, StockLevel, ConsolidatedStock, BranchStock
)
from src.domain.models.distribution import Transfer, DistributionResult
from src.infrastructure.persistence.mappers import StockMapper
from src.shared.utils.logging_utils import get_logger
from src.shared.utils.file_handler import ensure_directory_exists
from src.shared.constants import BRANCHES
from src.application.interfaces.repository import DataRepository
from src.core.validation.dates import extract_dates_from_header

logger = get_logger(__name__)


class PandasDataRepository(DataRepository):
    """
    Handles data persistence using Pandas and local CSV files.
    """

    def __init__(self, input_dir: str, output_dir: str, **kwargs):
        self._input_dir = input_dir
        self._output_dir = output_dir
        self._surplus_dir = kwargs.get('surplus_dir', output_dir)
        self._shortage_dir = kwargs.get('shortage_dir', output_dir)
        self._analytics_dir = kwargs.get('analytics_dir', output_dir)
        self._transfers_dir = kwargs.get('transfers_dir', output_dir)

    def load_branches(self) -> List[Branch]:
        return [Branch(name=name) for name in BRANCHES]

    def load_products(self) -> List[Product]:
        consolidated = self.load_consolidated_stock()
        return [item.product for item in consolidated]

    def _read_csv_with_header_detection(self, file_path: str) -> pd.DataFrame:
        """Reads CSV and handles optional 1-line date header using domain logic."""
        with open(file_path, 'r', encoding='utf-8-sig') as file_handle:
            first_line = file_handle.readline().strip()
        
        # Use domain logic for date header detection
        start_date, end_date = extract_dates_from_header(first_line)
        
        if start_date and end_date:
            logger.info(f"Detected date header in {file_path}, skipping first line.")
            return pd.read_csv(file_path, skiprows=1, encoding='utf-8-sig')
        
        # Fallback: Check for typical "Unnamed" artifacts if read incorrectly
        inventory_dataframe = pd.read_csv(file_path, encoding='utf-8-sig', nrows=5)
        if inventory_dataframe.columns[0].startswith('Unnamed') or 'الفترة من' in inventory_dataframe.columns[0]:
            logger.warning(f"Detected potential malformed header in {file_path}, retrying with skip.")
            return pd.read_csv(file_path, skiprows=1, encoding='utf-8-sig')

        return pd.read_csv(file_path, encoding='utf-8-sig')

    def load_consolidated_stock(self) -> List[ConsolidatedStock]:
        """Loads and maps consolidated stock from the latest input CSV."""
        if not os.path.exists(self._input_dir):
            return []
            
        files = [
            file_name for file_name in os.listdir(self._input_dir) 
            if file_name.endswith('.csv') and 'renamed' in file_name.lower()
        ]
        
        if not files:
            files = [file_name for file_name in os.listdir(self._input_dir) if file_name.endswith('.csv')]
            
        if not files:
            return []
            
        files.sort(
            key=lambda file_name: os.path.getmtime(os.path.join(self._input_dir, file_name)),
            reverse=True
        )
        
        csv_path = os.path.join(self._input_dir, files[0])
        num_days = 90 
        
        try:
            inventory_dataframe = self._read_csv_with_header_detection(csv_path)
            # Ensure column names are stripped of whitespace and BOM
            inventory_dataframe.columns = [column.strip().replace('\ufeff', '') for column in inventory_dataframe.columns]
            
            logger.info(f"Loaded CSV from {csv_path} with columns: {list(inventory_dataframe.columns)}")
            
            if 'code' not in inventory_dataframe.columns:
                logger.error(f"CRITICAL: 'code' column missing in {csv_path}!")
                # Try to find a column that looks like 'code'
                for column in inventory_dataframe.columns:
                    if 'code' in column.lower() or 'كود' in column:
                        logger.warning(f"Found potential match: '{column}'")

            results = []
            for _, row in inventory_dataframe.iterrows():
                stock_object = StockMapper.to_consolidated_stock(row, num_days)
                if stock_object:
                    results.append(stock_object)
            return results
        except Exception as error:
            logger.error(f"Error loading consolidated stock from {csv_path}: {error}")
            import traceback
            logger.error(traceback.format_exc())
            return []

    def save_branch_stocks(self, branch: Branch, stocks: List[BranchStock]) -> None:
        """Saves branch-specific stock levels to CSV."""
        inventory_dataframe = StockMapper.to_branch_dataframe(stocks)
        
        branch_dir = os.path.join(self._analytics_dir, branch.name)
        os.makedirs(branch_dir, exist_ok=True)
        
        output_path = os.path.join(branch_dir, f"main_analysis_{branch.name}.csv")
        inventory_dataframe.to_csv(output_path, index=False, encoding='utf-8-sig')

    def load_stock_levels(self, branch: Branch) -> Dict[str, StockLevel]:
        """Loads stock levels for a specific branch."""
        file_name = f"main_analysis_{branch.name}.csv"
        file_path = os.path.join(self._analytics_dir, branch.name, file_name)
        
        if not os.path.exists(file_path):
            logger.warning(f"Stock levels file not found for branch {branch.name}: {file_path}")
            return {}
            
        try:
            inventory_dataframe = pd.read_csv(file_path, encoding='utf-8-sig')
            stocks = {}
            for _, row in inventory_dataframe.iterrows():
                # Flexible column name lookup
                code_column = 'code' if 'code' in row else 'كود'
                if code_column in row:
                    stocks[str(row[code_column])] = StockMapper.to_stock_level(row)
            return stocks
        except Exception as error:
            print(f"Error loading stock levels for {branch.name}: {error}")
            return {}

    def save_transfers(self, transfers_list: List[Transfer]) -> None:
        """Saves a list of transfers to CSV files, grouped by source branch."""
        if not transfers_list:
            return
            
        # Group transfers by (source, target)
        branch_pairs = {}
        for transfer in transfers_list:
            key = (transfer.from_branch.name, transfer.to_branch.name)
            if key not in branch_pairs:
                branch_pairs[key] = []
            branch_pairs[key].append(transfer)
            
        os.makedirs(self._transfers_dir, exist_ok=True)
        
        for (source_name, target_name), transfers_list in branch_pairs.items():
            transfer_data = []
            for transfer in transfers_list:
                transfer_data.append({
                    'code': transfer.product.code,
                    'product_name': transfer.product.name,
                    'quantity_to_transfer': transfer.quantity,
                    'target_branch': target_name
                })
            
            transfer_dataframe = pd.DataFrame(transfer_data)
            transfer_dataframe = transfer_dataframe.sort_values(by='product_name')
            
            # Subdirectory for Step 11 compatibility
            subdirectory_name = f"transfers_from_{source_name}_to_other_branches"
            source_directory = os.path.join(self._output_dir, subdirectory_name)
            os.makedirs(source_directory, exist_ok=True)
            
            target_file_name = f"{source_name}_to_{target_name}.csv"
            output_file_path = os.path.join(source_directory, target_file_name)
            transfer_dataframe.to_csv(output_file_path, index=False, encoding='utf-8-sig')

    def save_remaining_surplus(self, results: List[DistributionResult]) -> None:
        """Saves products with remaining surplus to per-branch (total) and per-category files."""
        from src.core.domain.classification.product_classifier import classify_product_type
        from datetime import datetime
        
        timestamp = datetime.now().strftime("%Y%m%d")
        base_dir = self._surplus_dir
        if base_dir.endswith('csv'):
            base_dir = os.path.dirname(base_dir)

        # Group data by Branch -> Category
        grouped_data = {}
        for res in results:
            for branch_name, surplus in res.remaining_branch_surplus.items():
                if surplus > 0:
                    category = classify_product_type(res.product.name)
                    if branch_name not in grouped_data:
                        grouped_data[branch_name] = {}
                    if category not in grouped_data[branch_name]:
                        grouped_data[branch_name][category] = []
                    
                    grouped_data[branch_name][category].append({
                        'code': res.product.code,
                        'product_name': res.product.name,
                        'remaining_surplus': surplus
                    })

        for branch_name, categories in grouped_data.items():
            all_branch_items = []
            for category, items in categories.items():
                all_branch_items.extend(items)
                df = pd.DataFrame(items).sort_values('product_name', key=lambda x: x.str.lower())
                
                # CSV Output: base/csv/branch/surplus_branch_category_date.csv
                csv_dir = os.path.join(base_dir, "csv", branch_name)
                os.makedirs(csv_dir, exist_ok=True)
                csv_path = os.path.join(csv_dir, f"remaining_surplus_{branch_name}_{category}_{timestamp}.csv")
                df.to_csv(csv_path, index=False, encoding='utf-8-sig')
                
                # Excel Output
                excel_dir = os.path.join(base_dir, "excel", branch_name)
                os.makedirs(excel_dir, exist_ok=True)
                excel_path = os.path.join(excel_dir, f"remaining_surplus_{branch_name}_{category}_{timestamp}.xlsx")
                df.to_excel(excel_path, index=False)
            
            # Save Total Branch File (All Categories)
            if all_branch_items:
                df_total = pd.DataFrame(all_branch_items).sort_values('product_name', key=lambda x: x.str.lower())
                csv_path_total = os.path.join(base_dir, "csv", branch_name, f"remaining_surplus_{branch_name}_total_{timestamp}.csv")
                df_total.to_csv(csv_path_total, index=False, encoding='utf-8-sig')
                
                excel_path_total = os.path.join(base_dir, "excel", branch_name, f"remaining_surplus_{branch_name}_total_{timestamp}.xlsx")
                df_total.to_excel(excel_path_total, index=False)

    def save_shortage_report(self, results: List[DistributionResult]) -> None:
        """Saves products with a net shortage to total and category-split files."""
        from src.core.domain.classification.product_classifier import classify_product_type
        from datetime import datetime
        from src.shared.constants import BRANCHES
        from src.app.gui.utils.translations import BRANCH_NAMES, COLUMNS
        
        timestamp = datetime.now().strftime("%Y%m%d")
        base_dir = self._shortage_dir
        if base_dir.endswith('csv'):
            base_dir = os.path.dirname(base_dir)

        # Helper to format items with balances and sales
        def format_shortage_item(result: DistributionResult):
            item_data = {
                COLUMNS['code']: result.product.code,
                COLUMNS['product_name']: result.product.name,
                COLUMNS['shortage_quantity']: result.remaining_needed,
                COLUMNS['total_sales']: result.total_sales
            }
            # Add balance for each branch
            for branch_key in BRANCHES:
                column_name = f"رصيد {BRANCH_NAMES.get(branch_key, branch_key)}"
                item_data[column_name] = result.branch_balances.get(branch_key, 0.0) if result.branch_balances else 0.0
            return item_data

        # Group data by Category
        grouped_shortage_data = {}
        all_shortage_items = []
        for result in results:
            if result.remaining_needed > 0:
                product_category = classify_product_type(result.product.name)
                formatted_item = format_shortage_item(result)
                all_shortage_items.append(formatted_item)
                
                if product_category not in grouped_shortage_data:
                    grouped_shortage_data[product_category] = []
                grouped_shortage_data[product_category].append(formatted_item)

        for product_category, shortage_items in grouped_shortage_data.items():
            shortage_dataframe = pd.DataFrame(shortage_items).sort_values(COLUMNS['shortage_quantity'], ascending=False)
            
            # CSV Output
            csv_directory = os.path.join(base_dir, "csv")
            os.makedirs(csv_directory, exist_ok=True)
            csv_output_path = os.path.join(csv_directory, f"total_shortage_{product_category}_{timestamp}.csv")
            shortage_dataframe.to_csv(csv_output_path, index=False, encoding='utf-8-sig')
            
            # Excel Output
            excel_directory = os.path.join(base_dir, "excel")
            os.makedirs(excel_directory, exist_ok=True)
            excel_output_path = os.path.join(excel_directory, f"total_shortage_{product_category}_{timestamp}.xlsx")
            shortage_dataframe.to_excel(excel_output_path, index=False)

        # Save Consolidated Total File (All Categories)
        if all_shortage_items:
            total_shortage_dataframe = pd.DataFrame(all_shortage_items).sort_values(COLUMNS['shortage_quantity'], ascending=False)
            csv_total_path = os.path.join(base_dir, "csv", f"shortage_report_total_{timestamp}.csv")
            total_shortage_dataframe.to_csv(csv_total_path, index=False, encoding='utf-8-sig')
            
            excel_total_path = os.path.join(base_dir, "excel", f"shortage_report_total_{timestamp}.xlsx")
            total_shortage_dataframe.to_excel(excel_total_path, index=False)

    def load_transfers(self) -> List[Transfer]:
        """Loads transfers from the output directory (Step 7 output)."""
        all_transfers = []
        if not os.path.exists(self._output_dir):
            return []
            
        for root_directory, _, file_list in os.walk(self._output_dir):
            for file_name in file_list:
                # Expecting source_to_target.csv or similar
                if file_name.endswith('.csv') and '_to_' in file_name:
                    # Skip already split files (e.g. category files)
                    if any(product_category in file_name for product_category in ["tablets", "injections", "syrups", "creams", "sachets", "other"]):
                        continue
                        
                    file_full_path = os.path.join(root_directory, file_name)
                    base_filename = os.path.splitext(file_name)[0]
                    
                    # More robust parsing for source_to_target
                    # Format can be source_to_target or transfers_from_source_to_other_branches/...
                    filename_parts = base_filename.split('_to_')
                    if len(filename_parts) >= 2:
                        # Take the immediate parts around '_to_'
                        source_branch_name = filename_parts[0].split('_')[-1]
                        target_branch_name = filename_parts[1].split('_')[0]
                        
                        try:
                            transfer_dataframe = self._read_csv_with_header_detection(file_full_path)
                            for _, row_data in transfer_dataframe.iterrows():
                                all_transfers.append(Transfer(
                                    product=Product(code=str(row_data['code']), name=row_data['product_name']),
                                    from_branch=Branch(name=source_branch_name),
                                    to_branch=Branch(name=target_branch_name),
                                    quantity=int(row_data['quantity_to_transfer'])
                                ))
                        except Exception as loading_error:
                            print(f"Error loading transfers from {file_full_path}: {loading_error}")
        return all_transfers

    def save_split_transfers(self, transfers_list: List[Transfer], excel_directory: str) -> None:
        """Saves transfers split by product category into CSV and Excel."""
        from src.core.domain.classification.product_classifier import classify_product_type
        from datetime import datetime
        
        # Group by (sourceLabel, targetLabel, productCategory)
        grouped_transfers = {}
        for transfer_object in transfers_list:
            product_category = classify_product_type(transfer_object.product.name)
            group_key = (transfer_object.from_branch.name, transfer_object.to_branch.name, product_category)
            if group_key not in grouped_transfers:
                 grouped_transfers[group_key] = []
            grouped_transfers[group_key].append(transfer_object)
            
        timestamp_string = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        for (source_branch_name, target_branch_name, product_category), grouped_transfers_list in grouped_transfers.items():
            mapped_data = []
            for transfer_object in grouped_transfers_list:
                # Re-add product_type for Excel if needed (legacy had it sometimes)
                mapped_data.append({
                    'code': transfer_object.product.code,
                    'product_name': transfer_object.product.name,
                    'quantity_to_transfer': transfer_object.quantity,
                    'target_branch': target_branch_name
                })
            grouped_dataframe = pd.DataFrame(mapped_data).sort_values('product_name')
            
            # CSV: data/output/transfers/csv/source_to_target/source_to_target_timestamp_category.csv
            csv_subdirectory = os.path.join(self._output_dir, f"{source_branch_name}_to_{target_branch_name}")
            os.makedirs(csv_subdirectory, exist_ok=True)
            
            output_filename = f"{source_branch_name}_to_{target_branch_name}_{timestamp_string}_{product_category}"
            csv_output_path = os.path.join(csv_subdirectory, f"{output_filename}.csv")
            grouped_dataframe.to_csv(csv_output_path, index=False, encoding='utf-8-sig')
            
            # Excel: data/output/transfers/excel/transfers_excel_from_source_to_other_branches/source_to_target/source_to_target_timestamp_category.xlsx
            excel_subdirectory_path = os.path.join(
                excel_directory, 
                f"transfers_excel_from_{source_branch_name}_to_other_branches", 
                f"{source_branch_name}_to_{target_branch_name}"
            )
            os.makedirs(excel_subdirectory_path, exist_ok=True)
            excel_output_path = os.path.join(excel_subdirectory_path, f"{output_filename}.xlsx")
            
            try:
                with pd.ExcelWriter(excel_output_path, engine='openpyxl') as excel_writer:
                    grouped_dataframe.to_excel(excel_writer, index=False, sheet_name='Sheet1')
            except Exception as excel_saving_error:
                 print(f"Error saving Excel {excel_output_path}: {excel_saving_error}")

    def load_remaining_surplus(self, branch: Branch) -> List[Dict]:
        """Loads remaining surplus for a branch (Step 9 output)."""
        # Step 9 saves to surplus_dir/branch_name/remaining_surplus_branch.csv
        branch_directory_path = os.path.join(self._surplus_dir, branch.name)
        if not os.path.exists(branch_directory_path):
            return []
            
        surplus_data_list = []
        for file_name in os.listdir(branch_directory_path):
            if file_name.endswith('.csv') and 'remaining_surplus' in file_name:
                file_full_path = os.path.join(branch_directory_path, file_name)
                try:
                    surplus_dataframe = pd.read_csv(file_full_path, encoding='utf-8-sig')
                    for _, row_data in surplus_dataframe.iterrows():
                        surplus_data_list.append({
                            'code': str(row_data['code']),
                            'product_name': row_data['product_name'],
                            'quantity': int(row_data['remaining_surplus']),
                            'target_branch': 'admin',
                            'transfer_type': 'surplus'
                        })
                except Exception as surplus_loading_error:
                    print(f"Error loading surplus from {file_full_path}: {surplus_loading_error}")
        return surplus_data_list

    def save_combined_transfers(
        self, 
        branch: Branch,
        merged_data_list: List[Dict], 
        separate_data_list: List[Dict],
        timestamp_string: str
    ) -> None:
        """Saves combined transfers with Excel formatting."""
        # merged_data_list: list of {category: dataframe}
        # separate_data_list: list of {target: {category: dataframe}}
        
        base_output_directory = "data/output/combined_transfers"
        
        # Merged Output
        merged_csv_directory = os.path.join(base_output_directory, "merged", "csv", f"combined_transfers_from_{branch.name}_{timestamp_string}")
        merged_excel_directory = os.path.join(base_output_directory, "merged", "excel", f"combined_transfers_from_{branch.name}_{timestamp_string}")
        
        for merged_item in merged_data_list:
            product_category = merged_item['category']
            results_dataframe = merged_item['df']
            os.makedirs(merged_csv_directory, exist_ok=True)
            csv_output_path = os.path.join(merged_csv_directory, f"{branch.name}_combined_{product_category}.csv")
            results_dataframe.to_csv(csv_output_path, index=False, encoding='utf-8-sig')
            
            os.makedirs(merged_excel_directory, exist_ok=True)
            excel_output_path = os.path.join(merged_excel_directory, f"{branch.name}_combined_{product_category}.xlsx")
            self._save_formatted_excel(results_dataframe, excel_output_path)
            
        # Separate Output
        separate_csv_directory = os.path.join(base_output_directory, "separate", "csv", f"transfers_from_{branch.name}_{timestamp_string}")
        separate_excel_directory = os.path.join(base_output_directory, "separate", "excel", f"transfers_from_{branch.name}_{timestamp_string}")
        
        for separate_item in separate_data_list:
            target_branch_name = separate_item['target']
            product_category = separate_item['category']
            results_dataframe = separate_item['df']
            
            target_branch_csv_directory = os.path.join(separate_csv_directory, f"to_{target_branch_name}")
            os.makedirs(target_branch_csv_directory, exist_ok=True)
            csv_output_path = os.path.join(target_branch_csv_directory, f"transfer_from_{branch.name}_to_{target_branch_name}_{product_category}_{timestamp_string}.csv")
            results_dataframe.to_csv(csv_output_path, index=False, encoding='utf-8-sig')
            
            target_branch_excel_directory = os.path.join(separate_excel_directory, f"to_{target_branch_name}")
            os.makedirs(target_branch_excel_directory, exist_ok=True)
            excel_output_path = os.path.join(target_branch_excel_directory, f"transfer_from_{branch.name}_to_{target_branch_name}_{product_category}_{timestamp_string}.xlsx")
            self._save_formatted_excel(results_dataframe, excel_output_path)

    def _save_formatted_excel(self, results_dataframe: pd.DataFrame, file_path: str) -> None:
        """Saves DataFrame to Excel with Step 11 styling."""
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.formatting.rule import ColorScaleRule
        from openpyxl.utils import get_column_letter

        try:
            with pd.ExcelWriter(file_path, engine='openpyxl') as excel_writer:
                results_dataframe.to_excel(excel_writer, index=False, sheet_name='Sheet1')
                workbook_object = excel_writer.book
                worksheet_object = excel_writer.sheets['Sheet1']
                
                # Header Styling
                header_fill_style = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_font_style = Font(bold=True, color="FFFFFF")
                for header_cell in worksheet_object[1]:
                    header_cell.fill = header_fill_style
                    header_cell.font = header_font_style
                    header_cell.alignment = Alignment(horizontal='center')
                    
                # Column Widths
                column_width_mapping = {
                    'code': 12, 'product_name': 40, 'quantity_to_transfer': 15, 
                    'sender_balance': 15, 'receiver_balance': 15, 
                    'target_branch': 15, 'transfer_type': 15
                }
                for column_index, column_name in enumerate(results_dataframe.columns, 1):
                    column_letter = get_column_letter(column_index)
                    worksheet_object.column_dimensions[column_letter].width = column_width_mapping.get(column_name, 12)
                    
                # Borders
                thin_border_side = Side(style='thin')
                standard_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
                for table_row in worksheet_object.iter_rows(min_row=1, max_row=len(results_dataframe)+1, max_col=len(results_dataframe.columns)):
                    for row_cell in table_row:
                        row_cell.border = standard_border
                        
                # Color Scale for balances
                color_scale_rule = ColorScaleRule(
                    start_type='num', start_value=0, start_color='FF0000',
                    mid_type='num', mid_value=15, mid_color='FFFF00',
                    end_type='num', end_value=30, end_color='00FF00'
                )
                for column_index, column_name in enumerate(results_dataframe.columns, 1):
                    if column_name in ['sender_balance', 'receiver_balance']:
                        column_letter = get_column_letter(column_index)
                        worksheet_object.conditional_formatting.add(f"{column_letter}2:{column_letter}{len(results_dataframe)+1}", color_scale_rule)
                        
        except Exception as excel_formatting_error:
            print(f"Error saving formatted Excel {file_path}: {excel_formatting_error}")

    def list_outputs(self, category_name: str, branch_name_filter: Optional[str] = None) -> List[Dict]:
        """Lists available output artifacts for a given category and optional branch."""
        # mapping category to (base_directory, search_patterns)
        # search_patterns: {file_format: filename_prefix_pattern}
        output_category_mapping = {
            'transfers': {
                'base_directory': self._output_dir,
                'search_patterns': {'csv': '', 'excel': ''}
            },
            'surplus': {
                'base_directory': self._surplus_dir,
                'search_patterns': {'csv': '', 'excel': ''}
            },
            'shortage': {
                'base_directory': self._shortage_dir,
                'search_patterns': {'csv': '', 'excel': ''}
            },
            'merged': {
                'base_directory': os.path.join("data", "output", "combined_transfers", "merged"),
                'search_patterns': {'csv': 'combined_transfers_from_', 'excel': 'combined_transfers_from_'}
            },
            'separate': {
                'base_directory': os.path.join("data", "output", "combined_transfers", "separate"),
                'search_patterns': {'csv': 'transfers_from_', 'excel': 'transfers_from_'}
            }
        }
        
        if category_name not in output_category_mapping:
            return []
            
        category_config = output_category_mapping[category_name]
        base_search_directory = category_config['base_directory']
        output_artifact_results = []

        # Check for csv and excel subfolders
        for file_format in ['csv', 'excel']:
            format_directory = os.path.join(base_search_directory, file_format)
            if not os.path.exists(format_directory):
                # Fallback to base_search_directory if it's already a format directory (backward compatibility)
                if base_search_directory.endswith(file_format):
                    format_directory = base_search_directory
                else:
                    continue
            
            filename_prefix = category_config['search_patterns'].get(file_format, '')
            branch_sub_pattern = f"{filename_prefix}{branch_name_filter}" if branch_name_filter else filename_prefix
            
            # Special case for shortage (usually flat)
            if category_name == 'shortage':
                self._collect_files_recursive(format_directory, category_name, None, output_artifact_results)
                continue

            # Nested search for other categories
            for filesystem_item in os.listdir(format_directory):
                item_full_path = os.path.join(format_directory, filesystem_item)
                if not os.path.isdir(item_full_path):
                    continue
                    
                # Robust source branch matching for transfers
                has_branch_match = branch_sub_pattern in filesystem_item
                if category_name == 'transfers' and branch_name_filter:
                    has_branch_match = (
                        filesystem_item.startswith(branch_name_filter) or 
                        f"from_{branch_name_filter}" in filesystem_item or
                        f"From_{branch_name_filter}" in filesystem_item
                    )
                
                if has_branch_match:
                    # For 'separate', we have another level 'to_TARGET'
                    if category_name == 'separate':
                        for target_directory_name in os.listdir(item_full_path):
                            target_directory_full_path = os.path.join(item_full_path, target_directory_name)
                            if os.path.isdir(target_directory_full_path) and target_directory_name.startswith('to_'):
                                branch_metadata = branch_name_filter or filesystem_item.split('_')[-1]
                                self._collect_files_recursive(target_directory_full_path, category_name, branch_metadata, output_artifact_results)
                    else:
                        branch_metadata = branch_name_filter or filesystem_item
                        # For transfers, extract source branch from folder name if not provided
                        if category_name == 'transfers' and not branch_name_filter:
                            # pattern: transfers_from_SOURCE_TS/to_... or transfers_excel_from_SOURCE_TS/to_...
                            if 'from_' in filesystem_item and '_to_' in filesystem_item:
                                branch_metadata = filesystem_item.split('from_')[1].split('_to_')[0]
                        
                        self._collect_files_recursive(item_full_path, category_name, branch_metadata, output_artifact_results)
                        
        return output_artifact_results

    def _collect_files_recursive(self, search_directory: str, category_name: str, branch_identifier: Optional[str], collection_results: List[Dict]) -> None:
        """Helper to collect files recursively and add metadata."""
        if not os.path.exists(search_directory):
            return
            
        # Get the immediate folder name for zip paths
        parent_folder_name = os.path.basename(search_directory)
            
        for filesystem_item in os.listdir(search_directory):
            item_full_path = os.path.join(search_directory, filesystem_item)
            if os.path.isdir(item_full_path):
                # Recurse into subdirectory
                self._collect_files_recursive(item_full_path, category_name, branch_identifier, collection_results)
            elif os.path.isfile(item_full_path) and filesystem_item.endswith(('.csv', '.xlsx')):
                file_metadata = {
                    'name': filesystem_item,
                    'path': os.path.abspath(item_full_path),
                    'size': os.path.getsize(item_full_path),
                    'mtime': os.path.getmtime(item_full_path),
                    'category': category_name,
                    'branch': branch_identifier,
                    'folder_name': parent_folder_name
                }
                
                # Extract extra metadata for separate transfers
                if category_name == 'separate':
                    # search_directory format: .../csv/transfers_from_SOURCE_TS/to_TARGET
                    file_metadata['source_folder'] = os.path.basename(os.path.dirname(search_directory))
                    file_metadata['target_folder'] = parent_folder_name
                    
                    filename_without_extension = filesystem_item.split('.')[0]
                    if '_to_' in filename_without_extension:
                        filename_parts = filename_without_extension.split('_')
                        try:
                            to_label_index = filename_parts.index('to')
                            file_metadata['target_branch'] = filename_parts[to_label_index+1]
                            file_metadata['product_category'] = filename_parts[to_label_index+2]
                        except (ValueError, IndexError):
                            pass
                collection_results.append(file_metadata)
