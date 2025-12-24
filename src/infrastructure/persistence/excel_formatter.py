"""Excel formatting and styling logic for distribution reports."""

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter


def save_formatted_excel(dataframe: pd.DataFrame, file_path: str) -> None:
    """Saves DataFrame to Excel with specific styling and formatting."""
    try:
        with pd.ExcelWriter(file_path, engine='openpyxl') as excel_writer:
            dataframe.to_excel(excel_writer, index=False, sheet_name='Sheet1')
            worksheet = excel_writer.sheets['Sheet1']
            
            _apply_header_styles(worksheet)
            _apply_column_widths(worksheet, dataframe.columns)
            _apply_borders(worksheet, len(dataframe), len(dataframe.columns))
            _apply_conditional_formatting(worksheet, len(dataframe), dataframe.columns)
            
    except Exception as error:
        print(f"Error saving formatted Excel {file_path}: {error}")


def _apply_header_styles(worksheet) -> None:
    """Applies colors and alignment to the header row."""
    header_fill = PatternFill(
        start_color="4472C4", 
        end_color="4472C4", 
        fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')


def _apply_column_widths(worksheet, columns) -> None:
    """Sets standard column widths based on column names."""
    width_mapping = {
        'code': 12, 
        'product_name': 40, 
        'quantity_to_transfer': 15, 
        'sender_balance': 15, 
        'receiver_balance': 15, 
        'target_branch': 15, 
        'transfer_type': 15
    }
    
    for index, column_name in enumerate(columns, 1):
        column_letter = get_column_letter(index)
        width = width_mapping.get(column_name, 12)
        worksheet.column_dimensions[column_letter].width = width


def _apply_borders(worksheet, row_count: int, col_count: int) -> None:
    """Applies thin borders to data cells efficiently."""
    if row_count <= 0 or col_count <= 0:
        return
        
    thin = Side(style='thin')
    std_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    rows = worksheet.iter_rows(
        min_row=1, max_row=row_count + 1, max_col=col_count
    )
    for row in rows:
        for cell in row:
            cell.border = std_border


def _apply_conditional_formatting(worksheet, row_count: int, columns) -> None:
    """Applies color scale to balance columns."""
    color_scale = ColorScaleRule(
        start_type='num', start_value=0, start_color='FF0000',
        mid_type='num', mid_value=15, mid_color='FFFF00',
        end_type='num', end_value=30, end_color='00FF00'
    )
    
    balance_columns = ['sender_balance', 'receiver_balance']
    for index, column_name in enumerate(columns, 1):
        if column_name in balance_columns:
            column_letter = get_column_letter(index)
            range_string = f"{column_letter}2:{column_letter}{row_count + 1}"
            worksheet.conditional_formatting.add(range_string, color_scale)
