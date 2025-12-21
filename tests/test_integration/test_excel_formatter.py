"""Tests for Step 11 excel_formatter module"""

import os
import sys
from pathlib import Path

import pandas as pd
import pytest

# Add project root to path
PROJECT_ROOT = Path(__file__).parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT))


class TestExcelFormatting:
    """Tests for excel_formatter module"""
    
    def test_convert_to_excel_with_correct_key(self, tmp_path):
        """Test convert_to_excel_with_formatting with correct 'path' key"""
        from src.app.pipeline.step_11.excel_formatter import convert_to_excel_with_formatting
        
        # Create test CSV with proper structure
        csv_dir = tmp_path / "csv" / "combined_transfers_from_admin_20241220"
        csv_dir.mkdir(parents=True)
        
        df = pd.DataFrame({
            'code': ['001', '002', '003'],
            'product_name': ['Product A', 'Product B', 'Product C'],
            'quantity': [10, 20, 30],
            'target_branch': ['shahid', 'wardani', 'akba'],
            'transfer_type': ['normal', 'normal', 'surplus'],
            'sender_balance': [50, -5, 25],  # Negative to test red
            'receiver_balance': [20, 40, 10]
        })
        csv_path = csv_dir / "test_tablets_and_capsules.csv"
        df.to_csv(csv_path, index=False, encoding='utf-8-sig')
        
        excel_dir = tmp_path / "excel"
        excel_dir.mkdir()
        
        # Use correct 'path' key
        csv_files = [{
            'path': str(csv_path),
            'count': len(df)
        }]
        
        result = convert_to_excel_with_formatting(
            csv_files=csv_files,
            excel_output_dir=str(excel_dir)
        )
        
        assert isinstance(result, list)
        assert len(result) == 1
        assert result[0].endswith('.xlsx')
        assert os.path.exists(result[0])
    
    def test_convert_single_file_directly(self, tmp_path):
        """Test _convert_single_file function directly"""
        from src.app.pipeline.step_11.excel_formatter import _convert_single_file
        
        # Create test CSV
        csv_dir = tmp_path / "csv" / "transfers_from_shahid_20241220" / "to_admin"
        csv_dir.mkdir(parents=True)
        
        df = pd.DataFrame({
            'code': ['001', '002'],
            'product_name': ['Product A', 'Product B'],
            'quantity': [10, 20],
            'sender_balance': [50, 30],
            'receiver_balance': [20, 40]
        })
        csv_path = csv_dir / "test_file.csv"
        df.to_csv(csv_path, index=False, encoding='utf-8-sig')
        
        excel_dir = tmp_path / "excel"
        excel_dir.mkdir()
        
        result = _convert_single_file(str(csv_path), str(excel_dir))
        
        assert result is not None
        assert result.endswith('.xlsx')
        assert os.path.exists(result)
    
    def test_apply_conditional_formatting(self, tmp_path):
        """Test _apply_conditional_formatting function"""
        from src.app.pipeline.step_11.excel_formatter import _apply_conditional_formatting
        import openpyxl
        
        # Create test DataFrame
        df = pd.DataFrame({
            'code': ['001', '002', '003'],
            'product_name': ['A', 'B', 'C'],
            'quantity': [10, 20, 30],
            'sender_balance': [50, 0, 25],  # Include 0 for red color
            'receiver_balance': [20, 40, 5]
        })
        
        # Save to Excel first
        excel_path = tmp_path / "test.xlsx"
        df.to_excel(excel_path, index=False, engine='openpyxl')
        
        # Apply formatting
        _apply_conditional_formatting(str(excel_path), df)
        
        # Verify file still exists and is valid
        assert os.path.exists(excel_path)
        
        # Load and verify formatting was applied
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        
        # Check header styling
        assert ws['A1'].fill.start_color.rgb == "004472C4"  # Blue header
        
        # Check freeze panes
        assert ws.freeze_panes == 'A2'
        
        wb.close()
    
    def test_convert_empty_csv(self, tmp_path):
        """Test converting empty CSV file (headers only)"""
        from src.app.pipeline.step_11.excel_formatter import _convert_single_file
        
        csv_dir = tmp_path / "csv" / "folder"
        csv_dir.mkdir(parents=True)
        
        # Create CSV with headers but no data
        csv_path = csv_dir / "empty.csv"
        with open(csv_path, 'w', encoding='utf-8-sig') as f:
            f.write("code,product_name,sender_balance\n")
        
        excel_dir = tmp_path / "excel"
        excel_dir.mkdir()
        
        result = _convert_single_file(str(csv_path), str(excel_dir))
        
        # Should return None for empty file
        assert result is None
    
    def test_convert_nonexistent_file(self, tmp_path):
        """Test converting non-existent file"""
        from src.app.pipeline.step_11.excel_formatter import convert_to_excel_with_formatting
        
        excel_dir = tmp_path / "excel"
        excel_dir.mkdir()
        
        csv_files = [{
            'path': '/nonexistent/file.csv',
            'count': 0
        }]
        
        result = convert_to_excel_with_formatting(
            csv_files=csv_files,
            excel_output_dir=str(excel_dir)
        )
        
        # Should return empty list
        assert result == []
    
    def test_convert_with_missing_path_key(self, tmp_path):
        """Test converting with missing 'path' key"""
        from src.app.pipeline.step_11.excel_formatter import convert_to_excel_with_formatting
        
        excel_dir = tmp_path / "excel"
        excel_dir.mkdir()
        
        # Wrong key - should be 'path' not 'csv_path'
        csv_files = [{
            'csv_path': '/some/file.csv',
            'count': 0
        }]
        
        result = convert_to_excel_with_formatting(
            csv_files=csv_files,
            excel_output_dir=str(excel_dir)
        )
        
        # Should handle gracefully
        assert result == []
    
    def test_merged_folder_structure(self, tmp_path):
        """Test merged file folder structure (1 level)"""
        from src.app.pipeline.step_11.excel_formatter import _convert_single_file
        
        # Merged structure: csv/combined_transfers_from_xxx/file.csv
        csv_dir = tmp_path / "csv" / "combined_transfers_from_admin_20241220"
        csv_dir.mkdir(parents=True)
        
        df = pd.DataFrame({
            'code': ['001'],
            'product_name': ['A'],
            'sender_balance': [50],
            'receiver_balance': [20]
        })
        csv_path = csv_dir / "test.csv"
        df.to_csv(csv_path, index=False, encoding='utf-8-sig')
        
        excel_dir = tmp_path / "excel"
        excel_dir.mkdir()
        
        result = _convert_single_file(str(csv_path), str(excel_dir))
        
        assert result is not None
        # Check folder structure
        assert "combined_transfers_from_admin_20241220" in result
    
    def test_separate_folder_structure(self, tmp_path):
        """Test separate file folder structure (2 levels)"""
        from src.app.pipeline.step_11.excel_formatter import _convert_single_file
        
        # Separate structure: csv/transfers_from_xxx/to_yyy/file.csv
        csv_dir = tmp_path / "csv" / "transfers_from_admin_20241220" / "to_shahid"
        csv_dir.mkdir(parents=True)
        
        df = pd.DataFrame({
            'code': ['001'],
            'product_name': ['A'],
            'sender_balance': [50],
            'receiver_balance': [20]
        })
        csv_path = csv_dir / "test.csv"
        df.to_csv(csv_path, index=False, encoding='utf-8-sig')
        
        excel_dir = tmp_path / "excel"
        excel_dir.mkdir()
        
        result = _convert_single_file(str(csv_path), str(excel_dir))
        
        assert result is not None
        # Check folder structure includes both levels
        assert "to_shahid" in result
    
    def test_column_widths_applied(self, tmp_path):
        """Test that column widths are applied correctly"""
        from src.app.pipeline.step_11.excel_formatter import _apply_conditional_formatting
        import openpyxl
        
        df = pd.DataFrame({
            'code': ['001'],
            'product_name': ['A very long product name here'],
            'quantity': [10],
            'sender_balance': [50],
            'receiver_balance': [20],
            'target_branch': ['shahid'],
            'transfer_type': ['normal']
        })
        
        excel_path = tmp_path / "test.xlsx"
        df.to_excel(excel_path, index=False, engine='openpyxl')
        
        _apply_conditional_formatting(str(excel_path), df)
        
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        
        # Check product_name width is 40
        assert ws.column_dimensions['B'].width == 40
        
        # Check code width is 12
        assert ws.column_dimensions['A'].width == 12
        
        wb.close()
    
    def test_borders_applied(self, tmp_path):
        """Test that borders are applied to all cells"""
        from src.app.pipeline.step_11.excel_formatter import _apply_conditional_formatting
        import openpyxl
        
        df = pd.DataFrame({
            'code': ['001', '002'],
            'product_name': ['A', 'B'],
            'sender_balance': [50, 30]
        })
        
        excel_path = tmp_path / "test.xlsx"
        df.to_excel(excel_path, index=False, engine='openpyxl')
        
        _apply_conditional_formatting(str(excel_path), df)
        
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        
        # Check that cell A1 has border
        assert ws['A1'].border.left.style == 'thin'
        assert ws['A1'].border.right.style == 'thin'
        
        wb.close()
