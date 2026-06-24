"""Tests for infrastructure.excel.formatter module."""

import os
import tempfile
import shutil

import pandas as pd
import pytest
from openpyxl import load_workbook

from src.infrastructure.excel.formatter import (
    save_formatted_excel,
    _load_openpyxl_dependencies,
    _apply_header_styles,
    _apply_column_widths,
    _apply_borders,
    _apply_conditional_formatting,
)


@pytest.fixture
def sample_transfer_df():
    return pd.DataFrame({
        'code': ['001', '002'],
        'product_name': ['Product A', 'Product B'],
        'quantity_to_transfer': [10, 20],
        'target_branch': ['star', 'shahid'],
        'sender_balance': [5.0, 15.0],
        'receiver_balance': [2.0, 8.0],
    })


@pytest.fixture
def tmp_dir():
    d = tempfile.mkdtemp()
    yield d
    shutil.rmtree(d, ignore_errors=True)


class TestLoadOpenpyxlDependencies:
    def test_returns_all_keys(self):
        deps = _load_openpyxl_dependencies()
        expected_keys = [
            "PatternFill", "Font", "Alignment",
            "Border", "Side", "ColorScaleRule", "get_column_letter"
        ]
        for key in expected_keys:
            assert key in deps


class TestSaveFormattedExcel:
    def test_creates_file(self, sample_transfer_df, tmp_dir):
        path = os.path.join(tmp_dir, "out.xlsx")
        save_formatted_excel(sample_transfer_df, path)
        assert os.path.exists(path)

    def test_file_has_correct_data(self, sample_transfer_df, tmp_dir):
        path = os.path.join(tmp_dir, "out.xlsx")
        save_formatted_excel(sample_transfer_df, path)
        wb = load_workbook(path)
        ws = wb.active
        assert ws.cell(row=2, column=1).value == '001'
        assert ws.cell(row=2, column=2).value == 'Product A'
        assert ws.cell(row=3, column=3).value == 20

    def test_header_styling_applied(self, sample_transfer_df, tmp_dir):
        path = os.path.join(tmp_dir, "styled.xlsx")
        save_formatted_excel(sample_transfer_df, path)
        wb = load_workbook(path)
        ws = wb.active
        header_cell = ws.cell(row=1, column=1)
        assert header_cell.font.bold is True

    def test_empty_dataframe(self, tmp_dir):
        path = os.path.join(tmp_dir, "empty.xlsx")
        df = pd.DataFrame(columns=['code', 'product_name'])
        save_formatted_excel(df, path)
        assert os.path.exists(path)


class TestApplyBorders:
    def test_zero_rows_returns_early(self):
        deps = _load_openpyxl_dependencies()
        wb = __import__('openpyxl').Workbook()
        ws = wb.active
        _apply_borders(ws, 0, 3, deps)

    def test_zero_cols_returns_early(self):
        deps = _load_openpyxl_dependencies()
        wb = __import__('openpyxl').Workbook()
        ws = wb.active
        _apply_borders(ws, 3, 0, deps)


class TestApplyColumnWidths:
    def test_known_columns_get_custom_widths(self):
        deps = _load_openpyxl_dependencies()
        wb = __import__('openpyxl').Workbook()
        ws = wb.active
        columns = pd.Index(['code', 'product_name', 'quantity_to_transfer'])
        _apply_column_widths(ws, columns, deps)
        assert ws.column_dimensions['A'].width == 12
        assert ws.column_dimensions['B'].width == 40
        assert ws.column_dimensions['C'].width == 15

    def test_unknown_columns_get_default_width(self):
        deps = _load_openpyxl_dependencies()
        wb = __import__('openpyxl').Workbook()
        ws = wb.active
        columns = pd.Index(['unknown_col'])
        _apply_column_widths(ws, columns, deps)
        assert ws.column_dimensions['A'].width == 12


class TestApplyConditionalFormatting:
    def test_adds_rules_for_balance_columns(self):
        deps = _load_openpyxl_dependencies()
        wb = __import__('openpyxl').Workbook()
        ws = wb.active
        columns = pd.Index(['code', 'sender_balance', 'receiver_balance'])
        _apply_conditional_formatting(ws, 5, columns, deps)
        assert len(ws.conditional_formatting._cf_rules) == 2

    def test_no_rules_without_balance_columns(self):
        deps = _load_openpyxl_dependencies()
        wb = __import__('openpyxl').Workbook()
        ws = wb.active
        columns = pd.Index(['code', 'product_name'])
        _apply_conditional_formatting(ws, 5, columns, deps)
        assert len(ws.conditional_formatting._cf_rules) == 0
